# backend/docusense/local_vision.py
import fitz  # PyMuPDF
import easyocr
import pandas as pd
import re
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.conf import settings

# Initialize Reader
print("--- [Local Vision] Loading Layout-Aware OCR Model... ---")
reader = easyocr.Reader(['en'], gpu=True)

def extract_data_using_visual_layout(file_path):
    """
    Visual OCR Parser V2 (Refined):
    1. Clusters text into 'Visual Rows' (Y-axis).
    2. Separates Name from Marks by detecting the transition from Text to Digits.
    3. Identifies 'Total' as the distinct number before the Grade.
    """
    print(f"--- [Layout Engine] Processing: {os.path.basename(file_path)} ---")
    
    doc = fitz.open(file_path)
    all_rows = []
    
    # 1. OCR & ROW CLUSTERING
    for i, page in enumerate(doc):
        print(f"    -> Scanning Page {i+1}...")
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        img_bytes = pix.tobytes("png")
        results = reader.readtext(img_bytes, detail=1, paragraph=False)
        
        # Sort by Top-Y
        results.sort(key=lambda x: x[0][0][1])
        
        current_row = []
        last_y = -100
        
        for (bbox, text, conf) in results:
            y_top = bbox[0][1]
            x_left = bbox[0][0]
            if abs(y_top - last_y) < 15: # Row tolerance
                current_row.append({"text": text, "x": x_left})
            else:
                if current_row:
                    current_row.sort(key=lambda k: k['x'])
                    all_rows.append(current_row)
                current_row = [{"text": text, "x": x_left}]
                last_y = y_top
        if current_row:
            current_row.sort(key=lambda k: k['x'])
            all_rows.append(current_row)

    print(f"--- [Layout Engine] Detected {len(all_rows)} Visual Rows. Parsing... ---")

    # 2. PARSE ROWS INTO DATA
    parsed_students = []
    current_seat = "Unknown"
    current_pr = "-"
    current_name = "Unknown"
    current_sgpa = "0.00"
    
    prefixes = ["ECS", "CE", "ECOMP", "HM", "CV", "RAI", "ME", "EE", "ETC", "ITH", "AEC", "VAC", "SEC", "SHM", "HSS"]
    
    for row in all_rows:
        row_text_joined = " ".join([item['text'] for item in row])
        
        # A. Detect Header
        seat_match = re.search(r"Seat\s*No[:\.\s]*(\d+)", row_text_joined, re.IGNORECASE)
        if seat_match:
            current_seat = seat_match.group(1)
            # Find PR and Name in this or next rows? usually same row
            pr_match = re.search(r"PR\s*Number[:\.\s]*(\d+)", row_text_joined, re.IGNORECASE)
            if pr_match: current_pr = pr_match.group(1)
            name_match = re.search(r"Name[:\.\s]*([A-Z\s\.]+)", row_text_joined, re.IGNORECASE)
            if name_match: current_name = name_match.group(1).split("Paper")[0].strip()
            continue

        # B. Detect Footer
        sgpa_match = re.search(r"SGPA[:\.\s]*(\d+\.\d+)", row_text_joined, re.IGNORECASE)
        if sgpa_match: current_sgpa = sgpa_match.group(1)
        
        # C. Detect Subject Row
        if not row: continue
        first_word = row[0]['text'].upper().replace(" ", "").replace("-", "")
        
        is_subject = False
        for p in prefixes:
            if first_word.startswith(p) and any(char.isdigit() for char in first_word):
                is_subject = True
                break
        
        if is_subject:
            code = first_word
            remaining = row[1:]
            
            # --- INTELLIGENT SPLITTING ---
            name_parts = []
            numeric_parts = []
            
            # Find transition from Text Name -> Numeric Marks
            for idx, item in enumerate(remaining):
                txt = item['text']
                # Ignore garbage
                if "Paper" in txt: continue
                
                # If we hit a number or P/F status, the Name is done
                if (txt.isdigit()) or (txt in ['P', 'F'] and len(txt)==1) or (txt in ['A+', 'B+', 'A', 'B', 'C']):
                    # Add this and everything after to numeric parts
                    numeric_parts = [r['text'] for r in remaining[idx:]]
                    break
                else:
                    name_parts.append(txt)
            
            sub_name = " ".join(name_parts).strip()
            
            # Extract Marks Info from numeric_parts
            # Expectation: [Theory] [Status] ... [Total] [Grade]
            theory = "-"
            status = "-"
            total = "-"
            grade = "-"
            
            # Clean numeric parts
            # Filter out "Paper" again just in case
            numeric_parts = [x for x in numeric_parts if "Paper" not in x]
            
            if numeric_parts:
                # 1. Theory: Usually the first number
                for t in numeric_parts:
                    if t.isdigit():
                        theory = t
                        break
                
                # 2. Status: Look for 'P' or 'F' near the Theory mark
                for t in numeric_parts[:3]: # Check first 3 tokens
                    if t in ['P', 'F']:
                        status = t
                        break
                
                # 3. Grade: Usually the last item (letter)
                valid_grades = ['O', 'A+', 'A', 'B+', 'B', 'C', 'P', 'F']
                if numeric_parts[-1] in valid_grades:
                    grade = numeric_parts[-1]
                elif len(numeric_parts) > 1 and numeric_parts[-2] in valid_grades:
                    grade = numeric_parts[-2]
                    
                # 4. Total: The last number found in the list (before the grade)
                # Find all numbers
                nums = [n for n in numeric_parts if n.isdigit()]
                if nums:
                    # If multiple nums, Total is likely the biggest or last one
                    # But Theory is the first one. If there are 2 nums, second is Total.
                    if len(nums) >= 2:
                        total = nums[-1]
                    elif len(nums) == 1 and grade != "-":
                         # If only 1 num and we have a Grade, that num might be Total if it's high?
                         # Usually Theory is first. Let's assume Total is missing if only 1 num is found.
                         pass
            
            parsed_students.append({
                "Seat No": current_seat,
                "PR Number": current_pr,
                "Student Name": current_name,
                "Subject Code": code,
                "Subject Name": sub_name,
                "Theory": theory,
                "Status": status,
                "Total": total,
                "Grade": grade,
                "SGPA": current_sgpa # Placeholder, updated via post-processing or context
            })

    df = pd.DataFrame(parsed_students)
    
    # Backfill SGPA (Propagate SGPA up to the rows of the student)
    if not df.empty and 'SGPA' in df.columns:
        # We need to map Seat No to SGPA because SGPA appears at the end of the block
        # Group by Seat No and find the max SGPA (assuming non-zero)
        # Note: SGPA might be "0.00" initially for all rows, then found at bottom.
        # This simple logic assumes the SGPA was captured in 'current_sgpa' correctly by the end loop
        # But for rows emitted *before* the SGPA line, they have the old value.
        
        # Better: Create a Seat->SGPA map
        seat_sgpa_map = {}
        for row in all_rows:
            txt = " ".join([i['text'] for i in row])
            seat_m = re.search(r"Seat\s*No[:\.\s]*(\d+)", txt, re.IGNORECASE)
            if seat_m: curr = seat_m.group(1)
            sgpa_m = re.search(r"SGPA[:\.\s]*(\d+\.\d+)", txt, re.IGNORECASE)
            if sgpa_m and 'curr' in locals(): seat_sgpa_map[curr] = sgpa_m.group(1)
            
        df['SGPA'] = df['Seat No'].map(seat_sgpa_map).fillna("0.00")

    return df

def save_clean_excel(df, filename):
    if df.empty: return None
    safe_name = os.path.splitext(filename)[0]
    excel_name = f"{safe_name}_VISUAL_PARSED.xlsx"
    path = os.path.join(settings.MEDIA_ROOT, 'docusense_reports', excel_name)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    
    # Final Columns
    cols = ["Seat No", "PR Number", "Student Name", "Subject Code", "Subject Name", 
            "Theory", "Status", "Total", "Grade", "SGPA"]
    
    for c in cols:
        if c not in df.columns: df[c] = "-"
    df = df[cols]
    
    print(f"--- [Excel Writer] Saving to: {path} ---")
    df.to_excel(path, index=False)
    
    # Auto-Formatting
    wb = load_workbook(path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
    wb.save(path)
    return os.path.join('docusense_reports', excel_name)